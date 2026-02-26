"""
End-to-end API smoke tests for Project Buddy backend (reworked architecture).
Run from the backend/ directory:  python test_api.py
"""
import sys
from datetime import date, timedelta

import requests

BASE   = "http://localhost:5000/api"
passed = 0
failed = 0


def check(label, condition, detail=""):
    global passed, failed
    if condition:
        passed += 1
        print("  PASS  " + label)
    else:
        failed += 1
        print("  FAIL  " + label)
        if detail:
            print("        -> " + str(detail)[:120])


def j(r):
    try:
        return r.json()
    except Exception:
        return {}


# ---------------------------------------------------------------------------
# Pre-flight cleanup
_plans_r = requests.get(f"{BASE}/biweekly-plans")
if _plans_r.status_code == 200:
    for _p in _plans_r.json().get("data", {}).get("plans", []):
        if _p.get("name") == "Test Plan CI":
            requests.delete(f"{BASE}/biweekly-plans/{_p['id']}")

_proj_r = requests.get(f"{BASE}/projects")
if _proj_r.status_code == 200:
    for _p in _proj_r.json().get("data", {}).get("projects", []):
        if _p.get("name") in ("Tea Genome Analysis", "Ad-hoc: Meeting"):
            requests.delete(f"{BASE}/projects/{_p['id']}")

# ---------------------------------------------------------------------------
print("\n-- Health --")
r = requests.get("http://localhost:5000/health")
check("GET /health returns 200",   r.status_code == 200)
check("health.status == ok",       j(r).get("status") == "ok")

# ---------------------------------------------------------------------------
print("\n-- Projects (standalone) --")

r = requests.post(f"{BASE}/projects", json={
    "name":      "Tea Genome Analysis",
    "goal":      "Complete assembly and annotation",
    "color_tag": "#4472C4",
    "status":    "Active",
})
check("POST /projects -> 201",              r.status_code == 201, r.text)
proj_data = j(r).get("data", {})
proj_id   = proj_data.get("id")
check("Response contains project id",       isinstance(proj_id, int))
check("Project status = Active",            proj_data.get("status") == "Active")
check("No biweekly_plan_id in response",    "biweekly_plan_id" not in proj_data)

# List all projects
r = requests.get(f"{BASE}/projects")
check("GET /projects -> 200",              r.status_code == 200)
check("Project list has 1+ entries",       len(j(r)["data"]["projects"]) >= 1)

# Filter by status
r = requests.get(f"{BASE}/projects", params={"status": "Active"})
check("GET /projects?status=Active -> 200", r.status_code == 200)
found = any(p["id"] == proj_id for p in j(r)["data"]["projects"])
check("Our project appears in Active filter", found)

# Get by ID
r = requests.get(f"{BASE}/projects/{proj_id}")
check("GET /projects/{id} -> 200",         r.status_code == 200)
check("Project detail has activities list", "activities" in j(r)["data"])

# Update project
r = requests.put(f"{BASE}/projects/{proj_id}", json={"color_tag": "#FF5733"})
check("PUT /projects/{id} -> 200",          r.status_code == 200)
check("color_tag updated",                  j(r)["data"].get("color_tag") == "#FF5733")

# ---------------------------------------------------------------------------
print("\n-- Activities --")

r = requests.post(f"{BASE}/projects/{proj_id}/activities", json={
    "name":            "Illumina download SRP099527",
    "description":     "Download raw Illumina FASTQ files",
    "deliverables":    "Directory with raw FASTQ",
    "dependencies":    "Data quota available",
    "estimated_hours": 4.0,
})
check("POST /projects/{id}/activities -> 201",  r.status_code == 201, r.text)
act_data = j(r).get("data", {})
act_id   = act_data.get("id")
check("Response contains activity id",          isinstance(act_id, int))
check("Activity status = Not Started",          act_data.get("status") == "Not Started")
check("estimated_hours = 4.0",                  act_data.get("estimated_hours") == 4.0)
check("logged_hours = 0.0",                     act_data.get("logged_hours") == 0.0)

# Add a second activity
r2 = requests.post(f"{BASE}/projects/{proj_id}/activities", json={
    "name": "Filter Illumina Reads", "estimated_hours": 3.0
})
act_id2 = j(r2).get("data", {}).get("id")
check("Second activity created",               r2.status_code == 201)

# List activities
r = requests.get(f"{BASE}/projects/{proj_id}/activities")
check("GET /projects/{id}/activities -> 200",  r.status_code == 200)
check("Activities list has 2 entries",         len(j(r)["data"]["activities"]) == 2)

# Mark first activity complete
r = requests.put(f"{BASE}/activities/{act_id}", json={"status": "Complete"})
check("PUT /activities/{id} status=Complete",  r.status_code == 200)
check("Activity status = Complete",            j(r)["data"]["status"] == "Complete")

# ---------------------------------------------------------------------------
print("\n-- Biweekly Plans --")

monday = date.today() - timedelta(days=date.today().weekday())
start  = monday.isoformat()
end    = (monday + timedelta(days=9)).isoformat()

r = requests.post(f"{BASE}/biweekly-plans", json={
    "name":        "Test Plan CI",
    "description": "Automated smoke-test plan",
    "start_date":  start,
    "end_date":    end,
})
check("POST /biweekly-plans -> 201",          r.status_code == 201, r.text)
plan_data = j(r).get("data", {})
plan_id   = plan_data.get("id")
check("Response contains plan id",            isinstance(plan_id, int))
check("Plan status defaults to Active",       plan_data.get("status") == "Active")
check("No project_count, has sprint_activity_count",
      "sprint_activity_count" in plan_data and "project_count" not in plan_data)

# Duplicate name -> 409
r2 = requests.post(f"{BASE}/biweekly-plans", json={
    "name": "Test Plan CI", "start_date": start, "end_date": end
})
check("Duplicate plan name -> 409",           r2.status_code == 409)

# List plans
r = requests.get(f"{BASE}/biweekly-plans")
check("GET /biweekly-plans -> 200",           r.status_code == 200)
check("Plan list contains our plan",          any(p["id"] == plan_id for p in j(r)["data"]["plans"]))

# Get active
r = requests.get(f"{BASE}/biweekly-plans/active")
check("GET /biweekly-plans/active -> 200",    r.status_code == 200)
check("Active plan id matches",               j(r)["data"]["id"] == plan_id)

# Get by id
r = requests.get(f"{BASE}/biweekly-plans/{plan_id}")
check("GET /biweekly-plans/{id} -> 200",      r.status_code == 200)
check("Detail has sprint_activities list",    "sprint_activities" in j(r)["data"])
check("No projects list in plan detail",      "projects" not in j(r)["data"])

# Update
r = requests.put(f"{BASE}/biweekly-plans/{plan_id}", json={"description": "Updated desc"})
check("PUT /biweekly-plans/{id} -> 200",      r.status_code == 200)

# 404 on missing plan
r = requests.get(f"{BASE}/biweekly-plans/99999")
check("GET missing plan -> 404",              r.status_code == 404)

# ---------------------------------------------------------------------------
print("\n-- Sprint Activities --")

# Add activity to sprint
r = requests.post(f"{BASE}/biweekly-plans/{plan_id}/sprint-activities", json={
    "activity_id": act_id2,
    "notes":       "Focus on quality filtering this sprint",
})
check("POST /biweekly-plans/{id}/sprint-activities -> 201",   r.status_code == 201, r.text)
sa_data = j(r).get("data", {})
check("SprintActivity has activity_name",    sa_data.get("activity_name") == "Filter Illumina Reads")
check("SprintActivity has project_name",     sa_data.get("project_name") == "Tea Genome Analysis")

# Duplicate -> 409
r2 = requests.post(f"{BASE}/biweekly-plans/{plan_id}/sprint-activities", json={"activity_id": act_id2})
check("Duplicate sprint activity -> 409",    r2.status_code == 409)

# List sprint activities
r = requests.get(f"{BASE}/biweekly-plans/{plan_id}/sprint-activities")
check("GET /biweekly-plans/{id}/sprint-activities -> 200",   r.status_code == 200)
check("Sprint has 1 activity",               len(j(r)["data"]["sprint_activities"]) == 1)

# ---------------------------------------------------------------------------
print("\n-- Activity Logs --")

today_str = date.today().isoformat()
now_ts    = f"{today_str}T09:30:00+05:30"

# Log without plan_id (standalone)
r = requests.post(f"{BASE}/activity-logs", json={
    "project_id":  proj_id,
    "activity_id": act_id2,
    "comment":     "Started filtering Illumina reads, processed 500 samples",
    "duration_minutes": 60,
    "timestamp":   now_ts,
})
check("POST /activity-logs (no plan_id) -> 201",   r.status_code == 201, r.text)
log_data = j(r).get("data", {})
log_id   = log_data.get("id")
check("Response contains log id",                  isinstance(log_id, int))
check("project_name populated",                    log_data.get("project_name") == "Tea Genome Analysis")
check("biweekly_plan_id is null",                  log_data.get("biweekly_plan_id") is None)

# Log with plan_id
r2 = requests.post(f"{BASE}/activity-logs", json={
    "biweekly_plan_id": plan_id,
    "project_id":       proj_id,
    "activity_id":      act_id2,
    "comment":          "Continued filtering, 2000/5000 done",
    "duration_minutes": 45,
    "timestamp":        f"{today_str}T10:30:00+05:30",
})
log_id2 = j(r2).get("data", {}).get("id")
check("Second log (with plan_id) created",   r2.status_code == 201)

# Activity should now be In Progress
r = requests.get(f"{BASE}/projects/{proj_id}/activities")
acts = j(r)["data"]["activities"]
act2 = next((a for a in acts if a["id"] == act_id2), {})
check("Activity auto-set to In Progress",     act2.get("status") == "In Progress")
check("logged_hours = 1.75 (105 min)",        act2.get("logged_hours") == 1.75)

# List logs by date
r = requests.get(f"{BASE}/activity-logs", params={"date": today_str})
check("GET /activity-logs?date=today -> 200", r.status_code == 200)
check("2 logs returned",                      len(j(r)["data"]["logs"]) == 2)
check("total_hours = 1.75",                   j(r)["data"]["total_hours"] == 1.75)

# Edit a log
r = requests.put(f"{BASE}/activity-logs/{log_id}", json={"comment": "Updated comment"})
check("PUT /activity-logs/{id} -> 200",       r.status_code == 200)
check("Comment updated",                      j(r)["data"]["comment"] == "Updated comment")

# Validation: empty comment -> 422
r = requests.post(f"{BASE}/activity-logs", json={
    "project_id": proj_id, "comment": "", "timestamp": now_ts,
})
check("Empty comment -> 422",                 r.status_code == 422)

# Validation: duration out of range -> 422
r = requests.post(f"{BASE}/activity-logs", json={
    "project_id": proj_id, "comment": "x", "duration_minutes": 999, "timestamp": now_ts,
})
check("Duration > 480 -> 422",                r.status_code == 422)

# ---------------------------------------------------------------------------
print("\n-- Project Daily Notes --")

r = requests.post(f"{BASE}/project-notes", json={
    "project_id": proj_id,
    "date":       today_str,
    "what_i_did": "Completed FASTQ download and initial QC",
    "blockers":   "Network was slow in the morning",
    "next_steps": "Start Trimmomatic trimming tomorrow",
    "plan_id":    plan_id,
})
check("POST /project-notes -> 201",              r.status_code == 201, r.text)
note_data = j(r).get("data", {})
note_id   = note_data.get("id")
check("Note contains id",                        isinstance(note_id, int))
check("project_name populated",                  note_data.get("project_name") == "Tea Genome Analysis")

# Upsert: same project + date should update
r2 = requests.post(f"{BASE}/project-notes", json={
    "project_id": proj_id,
    "date":       today_str,
    "what_i_did": "Updated: completed 80% of QC",
})
check("Upsert (same date) -> 201",               r2.status_code == 201)
check("Same note id returned",                   j(r2)["data"]["id"] == note_id)
check("Content updated",                         "80%" in j(r2)["data"]["what_i_did"])

# List notes
r = requests.get(f"{BASE}/project-notes", params={"project_id": proj_id})
check("GET /project-notes?project_id -> 200",    r.status_code == 200)
check("1 note returned",                         len(j(r)["data"]["notes"]) == 1)

# Update note
r = requests.put(f"{BASE}/project-notes/{note_id}", json={"next_steps": "Run FastQC tomorrow"})
check("PUT /project-notes/{id} -> 200",          r.status_code == 200)
check("next_steps updated",                      j(r)["data"]["next_steps"] == "Run FastQC tomorrow")

# ---------------------------------------------------------------------------
print("\n-- Dashboard --")

r = requests.get(f"{BASE}/dashboard")
check("GET /dashboard -> 200",                r.status_code == 200, r.text[:200])
dash = j(r).get("data", {})
check("active_plan present",                  dash.get("active_plan") is not None)
check("active_plan.id matches",               dash.get("active_plan", {}).get("id") == plan_id)
check("days_remaining >= 0",                  dash.get("active_plan", {}).get("days_remaining", -1) >= 0)
check("sprint_activity_count present",        "sprint_activity_count" in dash.get("active_plan", {}))
check("No projects_count in active_plan",     "projects_count" not in dash.get("active_plan", {}))
check("projects list present (active only)",  isinstance(dash.get("projects"), list))
check("sprint_activities present",            isinstance(dash.get("sprint_activities"), list))
check("today_summary present",               dash.get("today_summary") is not None)
check("today total_hours_logged = 1.75",     dash.get("today_summary", {}).get("total_hours_logged") == 1.75)

# No DeepSeek summary yet
r = requests.get(f"{BASE}/dashboard/daily-summary", params={"date": today_str})
check("GET /dashboard/daily-summary (no data) -> 404", r.status_code == 404)

# ---------------------------------------------------------------------------
print("\n-- Excel Export --")

r = requests.get(f"{BASE}/biweekly-plans/{plan_id}/export-excel")
check("GET export-excel -> 200",              r.status_code == 200, r.text[:200])
check("Content-Type is XLSX",                 "spreadsheetml" in r.headers.get("Content-Type", ""))
check("Content-Disposition has attachment",   "attachment" in r.headers.get("Content-Disposition", ""))
check("Response body > 1 KB",                 len(r.content) > 1000)
check("Valid ZIP/XLSX magic bytes (PK)",      r.content[:2] == b"PK")

try:
    from io import BytesIO
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(r.content))
    sheet_names = wb.sheetnames
    check("Sheet 1: Overview exists",         "Overview" in sheet_names)
    check("Sheet 2: Projects & Activities",   "Projects & Activities" in sheet_names)
    check("Sheet 3: Time Tracking exists",    "Time Tracking" in sheet_names)
except Exception as exc:
    check("XLSX is readable by openpyxl",     False, str(exc))

# ---------------------------------------------------------------------------
print("\n-- Exports alias --")
r = requests.get(f"{BASE}/exports/plan-excel/{plan_id}")
check("GET /exports/plan-excel/{id} -> 200",  r.status_code == 200)
check("Alias also returns valid XLSX",         r.content[:2] == b"PK")

# ---------------------------------------------------------------------------
print("\n-- Remove sprint activity --")
r = requests.delete(f"{BASE}/biweekly-plans/{plan_id}/sprint-activities/{act_id2}")
check("DELETE sprint-activities/{act_id} -> 200",  r.status_code == 200)

r = requests.get(f"{BASE}/biweekly-plans/{plan_id}/sprint-activities")
check("Sprint is now empty",               len(j(r)["data"]["sprint_activities"]) == 0)

# ---------------------------------------------------------------------------
print("\n-- Delete / Cleanup --")

r = requests.delete(f"{BASE}/project-notes/{note_id}")
check("DELETE /project-notes/{id} -> 200",   r.status_code == 200)

r = requests.delete(f"{BASE}/activity-logs/{log_id2}")
check("DELETE /activity-logs/{id} -> 200",   r.status_code == 200)

r = requests.delete(f"{BASE}/activities/{act_id}")
check("DELETE /activities/{id} -> 200",      r.status_code == 200)

r = requests.delete(f"{BASE}/biweekly-plans/{plan_id}")
check("DELETE /biweekly-plans/{id} -> 200",  r.status_code == 200)

r = requests.get(f"{BASE}/biweekly-plans/{plan_id}")
check("Deleted plan -> 404",                 r.status_code == 404)

r = requests.get(f"{BASE}/biweekly-plans/active")
check("No active plan after delete -> 404",  r.status_code == 404)

r = requests.delete(f"{BASE}/projects/{proj_id}")
check("DELETE /projects/{id} -> 200",        r.status_code == 200)

r = requests.get(f"{BASE}/projects/{proj_id}")
check("Deleted project -> 404",              r.status_code == 404)

# ---------------------------------------------------------------------------
print("\n" + "=" * 60)
print(f"  Results: {passed} passed  |  {failed} failed")
print("=" * 60 + "\n")
sys.exit(1 if failed else 0)
