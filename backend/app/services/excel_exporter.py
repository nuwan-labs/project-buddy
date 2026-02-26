"""
Excel export service — generates styled .xlsx workbooks for biweekly plans.

Sheets produced:
  1. Overview       — plan metadata and sprint summary statistics
  2. Sprint Activities — breakdown by project/activity with date columns
  3. Time Tracking  — estimated vs. logged hours per sprint activity
"""
from __future__ import annotations

from collections import defaultdict
from datetime import date, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models import BiweeklyPlan

# ─────────────────────────────────────────────────────────────────────────────
# Style constants
# ─────────────────────────────────────────────────────────────────────────────

_BLUE_FILL  = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_LBLUE_FILL = PatternFill(start_color="D6E4F7", end_color="D6E4F7", fill_type="solid")
_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_YLOW_FILL  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_GREY_FILL  = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_RED_FILL   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

_WHITE_BOLD = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
_TITLE_FONT = Font(bold=True, name="Calibri", size=14)
_HEAD_FONT  = Font(bold=True, name="Calibri", size=11)
_BODY_FONT  = Font(name="Calibri", size=10)
_BOLD_BODY  = Font(bold=True, name="Calibri", size=10)

_CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_VCENTER = Alignment(horizontal="left",  vertical="center", wrap_text=True)

_DAY_ABBR = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]


def _thin_border() -> Border:
    s = Side(border_style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_cell(cell, fill=_BLUE_FILL, font=_WHITE_BOLD, align=_CENTER) -> None:
    cell.fill      = fill
    cell.font      = font
    cell.alignment = align
    cell.border    = _thin_border()


def _body_cell(cell, align=_LEFT, fill=None) -> None:
    cell.font      = _BODY_FONT
    cell.alignment = align
    cell.border    = _thin_border()
    if fill:
        cell.fill = fill


# ─────────────────────────────────────────────────────────────────────────────
# Date helpers
# ─────────────────────────────────────────────────────────────────────────────

def _get_workdays(start: date, end: date) -> list[date]:
    days: list[date] = []
    cur = start
    while cur <= end:
        if cur.weekday() < 5:
            days.append(cur)
        cur += timedelta(days=1)
    return days


def _week_groups(workdays: list[date]) -> list[tuple[int, int, str]]:
    if not workdays:
        return []
    groups: list[tuple[int, int, str]] = []
    cur_week   = workdays[0].isocalendar()[1]
    group_start = 0
    for i, d in enumerate(workdays):
        iso_week = d.isocalendar()[1]
        if iso_week != cur_week:
            groups.append((group_start, i - 1, f"Week of {workdays[group_start].strftime('%b %d')}"))
            group_start = i
            cur_week    = iso_week
    groups.append((group_start, len(workdays) - 1, f"Week of {workdays[group_start].strftime('%b %d')}"))
    return groups


# ─────────────────────────────────────────────────────────────────────────────
# Helper: group sprint activities by project
# ─────────────────────────────────────────────────────────────────────────────

def _sprint_projects(plan: BiweeklyPlan):
    """Return list of (project, [sprint_activities]) grouped by project."""
    by_project: dict[int, list] = defaultdict(list)
    project_map: dict[int, object] = {}
    for sa in plan.sprint_activities:
        if sa.activity and sa.activity.project:
            pid = sa.activity.project.id
            by_project[pid].append(sa)
            project_map[pid] = sa.activity.project
    return [(project_map[pid], acts) for pid, acts in by_project.items()]


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1 — Overview
# ─────────────────────────────────────────────────────────────────────────────

def _build_overview(ws, plan: BiweeklyPlan) -> None:
    from datetime import datetime

    ws.title = "Overview"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 42

    ws.merge_cells("A1:B1")
    ws["A1"] = f"Sprint Plan: {plan.name}"
    ws["A1"].font      = _TITLE_FONT
    ws["A1"].alignment = _LEFT
    ws["A1"].fill      = _LBLUE_FILL
    ws.row_dimensions[1].height = 28

    meta = [
        ("Period",      f"{plan.start_date}  →  {plan.end_date}"),
        ("Status",      plan.status),
        ("Description", plan.description or "—"),
        ("Generated",   datetime.now().strftime("%B %d, %Y at %I:%M %p")),
    ]
    for i, (label, value) in enumerate(meta, start=2):
        a_cell = ws.cell(row=i, column=1, value=label)
        b_cell = ws.cell(row=i, column=2, value=value)
        a_cell.font   = _HEAD_FONT
        a_cell.fill   = _LBLUE_FILL
        a_cell.border = _thin_border()
        b_cell.font   = _BODY_FONT
        b_cell.border = _thin_border()

    sprint_acts = plan.sprint_activities
    total_acts  = len(sprint_acts)
    completed   = sum(1 for sa in sprint_acts if sa.activity and sa.activity.status == "Complete")
    in_progress = sum(1 for sa in sprint_acts if sa.activity and sa.activity.status == "In Progress")
    pct         = f"{completed / total_acts * 100:.1f}%" if total_acts else "0%"
    proj_groups = _sprint_projects(plan)

    stats_start = 7
    ws.merge_cells(f"A{stats_start}:B{stats_start}")
    ws[f"A{stats_start}"] = "Sprint Statistics"
    _header_cell(ws[f"A{stats_start}"])

    stats = [
        ("Projects in Sprint",    len(proj_groups)),
        ("Sprint Activities",     total_acts),
        ("Completed",             completed),
        ("In Progress",           in_progress),
        ("Not Started",           total_acts - completed - in_progress),
        ("Overall Completion",    pct),
    ]
    for i, (label, value) in enumerate(stats, start=stats_start + 1):
        a = ws.cell(row=i, column=1, value=label)
        b = ws.cell(row=i, column=2, value=value)
        a.font   = _HEAD_FONT
        a.border = _thin_border()
        b.font   = _BODY_FONT
        b.border = _thin_border()
        if label == "Completed":
            b.fill = _GREEN_FILL
        elif label == "In Progress":
            b.fill = _YLOW_FILL

    proj_start = stats_start + len(stats) + 2
    ws.merge_cells(f"A{proj_start}:B{proj_start}")
    ws[f"A{proj_start}"] = "Sprint Projects"
    _header_cell(ws[f"A{proj_start}"])

    proj_headers = ["Project Name", "Activities in Sprint"]
    for col, h in enumerate(proj_headers, start=1):
        cell = ws.cell(row=proj_start + 1, column=col, value=h)
        _header_cell(cell, fill=_LBLUE_FILL, font=_HEAD_FONT)

    for i, (proj, acts) in enumerate(proj_groups, start=proj_start + 2):
        done = sum(1 for sa in acts if sa.activity and sa.activity.status == "Complete")
        total = len(acts)
        pct_p = f"{done / total * 100:.0f}%" if total else "0%"
        a = ws.cell(row=i, column=1, value=proj.name)
        b = ws.cell(row=i, column=2, value=f"{done}/{total} activities  ({pct_p})")
        _body_cell(a)
        _body_cell(b)
        if done == total and total:
            b.fill = _GREEN_FILL
        elif done:
            b.fill = _YLOW_FILL


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2 — Sprint Activities
# ─────────────────────────────────────────────────────────────────────────────

def _build_activities(ws, plan: BiweeklyPlan) -> None:
    ws.title = "Projects & Activities"

    try:
        start_d = date.fromisoformat(plan.start_date)
        end_d   = date.fromisoformat(plan.end_date)
    except ValueError:
        start_d = end_d = date.today()

    workdays = _get_workdays(start_d, end_d)
    groups   = _week_groups(workdays)
    n_fixed  = 6   # Project | Activity | Deliverables | Dependencies | Est.Hrs | Status

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_fixed + len(workdays))
    ws.cell(row=1, column=1, value=f"{plan.name}  |  {plan.start_date} → {plan.end_date}").font = _TITLE_FONT
    ws.cell(row=1, column=1).alignment = _LEFT
    ws.cell(row=1, column=1).fill = _LBLUE_FILL
    ws.row_dimensions[1].height = 24

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_fixed)
    _header_cell(ws.cell(row=2, column=1, value="Activity Details"))

    for g_start, g_end, label in groups:
        sc = n_fixed + g_start + 1
        ec = n_fixed + g_end  + 1
        if sc <= ec:
            ws.merge_cells(start_row=2, start_column=sc, end_row=2, end_column=ec)
        cell = ws.cell(row=2, column=sc, value=label)
        _header_cell(cell, fill=_LBLUE_FILL, font=_HEAD_FONT)

    fixed_headers = ["Project", "Activity", "Deliverables", "Dependencies", "Est. Hrs", "Status"]
    for col, h in enumerate(fixed_headers, start=1):
        _header_cell(ws.cell(row=3, column=col, value=h))

    for idx, d in enumerate(workdays):
        label = f"{_DAY_ABBR[d.weekday()]}\n{d.strftime('%d')}"
        _header_cell(ws.cell(row=3, column=n_fixed + idx + 1, value=label))
    ws.row_dimensions[3].height = 30

    data_row = 4
    for proj, sprint_acts in _sprint_projects(plan):
        proj_start_row = data_row
        for act_idx, sa in enumerate(sprint_acts):
            act = sa.activity
            proj_cell = ws.cell(row=data_row, column=1,
                                value=proj.name if act_idx == 0 else "")
            _body_cell(proj_cell, fill=_GREY_FILL)
            if act_idx == 0:
                proj_cell.font = _BOLD_BODY

            _body_cell(ws.cell(row=data_row, column=2, value=act.name if act else ""))
            _body_cell(ws.cell(row=data_row, column=3, value=(act.deliverables or "") if act else ""))
            _body_cell(ws.cell(row=data_row, column=4, value=(act.dependencies or "") if act else ""))
            _body_cell(ws.cell(row=data_row, column=5, value=(act.estimated_hours or 0) if act else 0), align=_CENTER)

            act_status = act.status if act else "—"
            status_cell = ws.cell(row=data_row, column=6, value=act_status)
            status_fill = {
                "Complete":    _GREEN_FILL,
                "In Progress": _YLOW_FILL,
                "Blocked":     _RED_FILL,
            }.get(act_status)
            _body_cell(status_cell, align=_CENTER, fill=status_fill)

            for idx in range(len(workdays)):
                _body_cell(ws.cell(row=data_row, column=n_fixed + idx + 1, value=""), align=_CENTER)

            data_row += 1

        if len(sprint_acts) > 1:
            ws.merge_cells(
                start_row=proj_start_row, start_column=1,
                end_row=data_row - 1,     end_column=1,
            )
            ws.cell(proj_start_row, 1).alignment = _CENTER

    fixed_widths = {1: 22, 2: 30, 3: 22, 4: 18, 5: 8, 6: 13}
    for col, w in fixed_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    for idx in range(len(workdays)):
        ws.column_dimensions[get_column_letter(n_fixed + idx + 1)].width = 6

    ws.freeze_panes = "C4"


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 3 — Time Tracking
# ─────────────────────────────────────────────────────────────────────────────

def _build_time_tracking(ws, plan: BiweeklyPlan, db) -> None:
    ws.title = "Time Tracking"

    from sqlalchemy import func
    from app.models import ActivityLog

    agg = (
        db.query(
            ActivityLog.project_id,
            ActivityLog.activity_id,
            func.sum(ActivityLog.duration_minutes).label("total_min"),
        )
        .filter(ActivityLog.biweekly_plan_id == plan.id)
        .group_by(ActivityLog.project_id, ActivityLog.activity_id)
        .all()
    )
    hours_lookup: dict[tuple[int, int | None], float] = {
        (r.project_id, r.activity_id): round((r.total_min or 0) / 60, 2)
        for r in agg
    }

    headers = ["Project", "Activity", "Est. Hours", "Logged Hours", "Remaining", "% Done", "Status"]
    for col, h in enumerate(headers, start=1):
        _header_cell(ws.cell(row=1, column=col, value=h))
    ws.row_dimensions[1].height = 20

    data_row = 2
    for proj, sprint_acts in _sprint_projects(plan):
        proj_start_row = data_row
        for act_idx, sa in enumerate(sprint_acts):
            act       = sa.activity
            logged    = hours_lookup.get((proj.id, act.id if act else None), 0.0)
            est       = (act.estimated_hours or 0.0) if act else 0.0
            remaining = max(round(est - logged, 2), 0.0)
            pct       = f"{logged / est * 100:.0f}%" if est else "N/A"
            act_status = act.status if act else "—"

            _body_cell(ws.cell(data_row, 1, value=proj.name if act_idx == 0 else ""),
                       fill=_GREY_FILL)
            if act_idx == 0:
                ws.cell(data_row, 1).font = _BOLD_BODY

            _body_cell(ws.cell(data_row, 2, value=act.name if act else ""))
            _body_cell(ws.cell(data_row, 3, value=est),       align=_CENTER)

            logged_cell = ws.cell(data_row, 4, value=logged)
            log_fill = _GREEN_FILL if act_status == "Complete" else (_YLOW_FILL if logged > 0 else None)
            _body_cell(logged_cell, align=_CENTER, fill=log_fill)

            _body_cell(ws.cell(data_row, 5, value=remaining), align=_CENTER)
            _body_cell(ws.cell(data_row, 6, value=pct),       align=_CENTER)

            status_fill = {
                "Complete":    _GREEN_FILL,
                "In Progress": _YLOW_FILL,
                "Blocked":     _RED_FILL,
            }.get(act_status)
            _body_cell(ws.cell(data_row, 7, value=act_status), align=_CENTER, fill=status_fill)

            data_row += 1

        if len(sprint_acts) > 1:
            ws.merge_cells(
                start_row=proj_start_row, start_column=1,
                end_row=data_row - 1,     end_column=1,
            )
            ws.cell(proj_start_row, 1).alignment = _CENTER

    for col, w in {1: 22, 2: 30, 3: 10, 4: 13, 5: 10, 6: 8, 7: 13}.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "C2"


# ─────────────────────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────────────────────

def generate_biweekly_plan_excel(plan: BiweeklyPlan, db=None) -> BytesIO:
    """Generate a styled XLSX workbook for a biweekly sprint plan."""
    wb = Workbook()

    _build_overview(wb.active, plan)

    ws2 = wb.create_sheet("Projects & Activities")
    _build_activities(ws2, plan)

    if db is not None:
        ws3 = wb.create_sheet("Time Tracking")
        _build_time_tracking(ws3, plan, db)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
